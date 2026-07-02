"""向「三工位-15台测试.xlsx」填入三个模块的测试数据"""
import openpyxl

path = r'C:\Users\20210817\Desktop\三工位-15台测试.xlsx'
wb = openpyxl.load_workbook(path)

# ====== 外购件BOM表 ======
# Row 1: title | Row 2: meta | Row 3: headers | Row 4+: data
# Cols: A序号 B型号 C名称规格 D品牌 E单台数量 F装机台数n G=n台数量 H单位 I货期 J单价 K总价 L提交人 M验收人 W供应商链接(23)
ws1 = wb['外购件BOM表']
# 装机台数 n = 1 (row 2, col J says "本次组装机台数量 n=1")

parts = [
    # [型号, 名称规格, 品牌, 单台数量, 单位, 货期(天), 单价, 提交人, 供应商链接]
    # 轴承/导轨/丝杆
    ['6204-2RS', '深沟球轴承 6204-2RS 内径20mm', 'SKF', 4, '个', 7, 35, '熊振', ''],
    ['6005-2RS', '深沟球轴承 6005-2RS 内径25mm', 'NSK', 4, '个', 7, 42, '熊振', ''],
    ['HGH20CA-1000', '直线导轨 HGH20CA 长度1000mm 含滑块×2', '上银HIWIN', 2, '套', 14, 1280, '熊振', ''],
    ['SFU1605-1000', '滚珠丝杆 SFU1605 导程5mm 长度1000mm', '上银HIWIN', 2, '套', 14, 980, '熊振', ''],
    ['DS40-AL', '联轴器 DS40 梅花型 铝合金 夹紧式', '米思米', 4, '个', 5, 45, '熊振', ''],
    # 气缸/气动
    ['MDBB32-100Z', '气缸 MDBB32-100Z 双作用 32mm×100mm', 'SMC', 6, '个', 10, 320, '熊振', ''],
    ['MDBB40-150Z', '气缸 MDBB40-150Z 双作用 40mm×150mm', 'SMC', 3, '个', 10, 420, '熊振', ''],
    ['D-M9B', '气缸传感器 D-M9B 无触点 固态开关', 'SMC', 12, '个', 5, 55, '熊振', ''],
    ['AS2201F-01-08S', '调速阀 AS2201F-01-08S 单向节流', 'SMC', 15, '个', 3, 18, '熊振', ''],
    ['ZP3-T08BUN-J6-B5', '真空吸盘 ZP3-T08BUN 丁腈橡胶 平型', 'SMC', 12, '个', 5, 25, '熊振', ''],
    ['ZH05BS-02-02', '真空发生器 ZH05BS 喷嘴直径0.5mm', 'SMC', 4, '个', 7, 180, '熊振', ''],
    ['KQ2H06-01AS', '气管直通接头 KQ2H06-01AS φ6-1/8', 'SMC', 40, '个', 3, 3.5, '熊振', ''],
    ['KQ2L06-01AS', '气管弯头 KQ2L06-01AS φ6-1/8', 'SMC', 30, '个', 3, 4.5, '熊振', ''],
    # 拖链/缓冲器
    ['2500.038.100.0', '拖链 E2/000 25×38 桥式 内高25mm', '易格斯IGUS', 4, '根', 10, 380, '熊振', ''],
    ['RB1007S', '油压缓冲器 RB1007S 行程7mm', 'SMC', 8, '个', 5, 65, '熊振', ''],
    # 紧固件
    ['GB70-M6×20-12.9', '内六角圆柱头螺栓 M6×20 12.9级 发黑', '国产', 200, '个', 2, 0.15, '熊振', ''],
    ['GB70-M8×25-12.9', '内六角圆柱头螺栓 M8×25 12.9级 发黑', '国产', 150, '个', 2, 0.25, '熊振', ''],
    ['GB93-M6', '弹簧垫圈 M6 65Mn 发黑', '国产', 200, '个', 2, 0.08, '熊振', ''],
    ['GB97-M6', '平垫圈 M6 A2不锈钢', '国产', 200, '个', 2, 0.06, '熊振', ''],
    # 定位/密封
    ['MSH8-30', '定位销 D8×30 淬火钢 圆柱型', '米思米', 16, '个', 3, 3.5, '熊振', ''],
    ['NBR-O30', 'O型圈套装 NBR 丁腈橡胶 30种规格', '国产', 2, '盒', 3, 45, '熊振', ''],
    ['PTFE-2mm', '密封垫片 PTFE 聚四氟乙烯 2mm厚', '国产', 4, '张', 5, 25, '熊振', ''],
    # 润滑/辅料
    ['MOBIL-EP2-2KG', '润滑脂 美孚EP2 复合锂基 2kg/桶', '美孚', 2, '桶', 3, 68, '熊振', ''],
    ['LOCTITE-242-50ML', '螺纹胶 乐泰242 中强度 50ml', '乐泰', 3, '支', 3, 28, '熊振', ''],
    ['WD40-400ML', '防锈清洁剂 WD-40 多用途 400ml', 'WD-40', 5, '瓶', 2, 18, '熊振', ''],
]

start_row = 5
n_val = 1  # 装机台数

for i, row in enumerate(parts):
    r = start_row + i
    model, name, brand, qty, unit, lead_time, price, submitter, link = row
    ws1.cell(row=r, column=1, value=i + 1)           # 序号
    ws1.cell(row=r, column=2, value=model)             # 型号
    ws1.cell(row=r, column=3, value=name)              # 名称规格
    ws1.cell(row=r, column=4, value=brand)             # 品牌
    ws1.cell(row=r, column=5, value=qty)               # 单台数量
    ws1.cell(row=r, column=6, value=n_val)             # 装机台数 n
    ws1.cell(row=r, column=7, value=f'=E{r}*F{r}')    # n台数量
    ws1.cell(row=r, column=8, value=unit)              # 单位
    ws1.cell(row=r, column=9, value=lead_time)         # 货期
    ws1.cell(row=r, column=10, value=price)            # 单价
    ws1.cell(row=r, column=11, value=f'=J{r}*G{r}')   # 总价
    ws1.cell(row=r, column=12, value=submitter)        # 提交人
    ws1.cell(row=r, column=13, value='王工')           # 验收人
    if link:
        ws1.cell(row=r, column=23, value=link)         # 供应商链接

print(f'外购件BOM表: 写入 {len(parts)} 行 (Row {start_row}-{start_row+len(parts)-1})')

# ====== 外加工钣金亚克力件BOM件 ======
# Cols: A序号 B图号 C零件名称 D供应商 E材质 F单位 G每台数量 H装机台数 I=N台数量 J表面处理 K单价 L总金额 M备注
ws2 = wb['外加工钣金亚克力件BOM件']

fab_parts = [
    # [图号, 零件名称, 供应商, 材质, 单位, 每台数量, 表面处理, 单价]
    ['SWG-BP-001', '钣金底板 1200×800×20mm', '外协加工-钣金', 'Q235A 冷轧钢板', '块', 1, '喷塑 RAL7035', 850],
    ['SWG-CL-001', '钣金立柱 方管100×100×6×800mm', '外协加工-钣金', 'Q235A 方管', '根', 4, '喷塑 RAL7035', 320],
    ['SWG-BM-001', '钣金横梁 方管80×80×5×1000mm', '外协加工-钣金', 'Q235A 方管', '根', 6, '喷塑 RAL7035', 280],
    ['SWG-MP-001', '电机安装板 300×200×12mm', '外协加工-钣金', 'Q235A 热轧板', '块', 6, '镀锌', 120],
    ['SWG-SB-001', '传感器支架 5mm厚 折弯件', '外协加工-钣金', '304不锈钢', '个', 12, '拉丝处理', 45],
    ['SWG-CA-001', '气缸安装座 L型 焊接件', '外协加工-焊接', 'Q235A', '个', 9, '喷塑 RAL7035', 65],
    ['SWG-GB-001', '导轨垫块 20×40×200mm', '外协加工-机加', '45#钢 调质', '块', 8, '发黑', 55],
    ['SWG-BH-001', '皮带轮护罩 1.5mm钣金折弯', '外协加工-钣金', '304不锈钢', '个', 4, '拉丝', 150],
    ['SWG-CV-001', '走线槽盖板 1.2mm SPCC', '外协加工-钣金', 'SPCC', '块', 8, '喷塑 RAL7035', 35],
    ['SWG-JB-001', '接线盒 200×150×100mm', '外协加工-钣金', 'Q235A 1.5mm', '个', 3, '喷塑 RAL7035', 95],
    ['SWG-ES-001', '急停按钮安装盒 铝合金', '外协加工-机加', 'ADC12 压铸铝', '个', 2, '喷砂氧化', 55],
    # 亚克力/非金属
    ['SWG-AC-001', '亚克力防护外罩 5mm透明', '外协加工-亚克力', 'PMMA 亚克力 透明', '套', 1, '激光切割+折弯', 1800],
    ['SWG-AV-001', '亚克力观察窗 8mm透明', '外协加工-亚克力', 'PMMA 亚克力 透明', '块', 6, '抛光边缘', 85],
    ['SWG-AL-001', '铝合金面板 6061-T6 6mm', '外协加工-机加', '6061-T6 铝合金', '块', 1, '喷砂氧化', 380],
    ['SWG-NS-001', '尼龙滑块 30×40×20mm', '外协加工-机加', 'PA66 尼龙', '个', 16, 'CNC精加工', 25],
]

start_row2 = 5
for i, row in enumerate(fab_parts):
    r = start_row2 + i
    drawing, name, supplier, material, unit, qty, surface, price = row
    ws2.cell(row=r, column=1, value=i + 1)             # 序号
    ws2.cell(row=r, column=2, value=drawing)             # 图号
    ws2.cell(row=r, column=3, value=name)                # 零件名称
    ws2.cell(row=r, column=4, value=supplier)            # 供应商
    ws2.cell(row=r, column=5, value=material)            # 材质
    ws2.cell(row=r, column=6, value=unit)                # 单位
    ws2.cell(row=r, column=7, value=qty)                 # 每台数量
    ws2.cell(row=r, column=8, value=n_val)               # 装机台数
    ws2.cell(row=r, column=9, value=f'=G{r}*H{r}')      # N台数量
    ws2.cell(row=r, column=10, value=surface)            # 表面处理
    ws2.cell(row=r, column=11, value=price)              # 单价
    ws2.cell(row=r, column=12, value=f'=K{r}*I{r}')     # 总金额

print(f'外加工钣金亚克力件BOM件: 写入 {len(fab_parts)} 行')

# ====== 视觉BOM ======
# Cols: A序号 B型号 C名称规格 D品牌 E单台数量 F装机台数n G=n台数量 H单位 I单价 J总价 K提交人 L验收人 V供应商链接(22)
ws3 = wb['视觉BOM']

vis_parts = [
    # [型号, 名称规格, 品牌, 单台数量, 单位, 单价, 提交人]
    ['MER-503-20GM-P', '工业相机 500万像素 黑白 GigE接口 2/3" CMOS', '大恒图像', 3, '台', 2850, '熊振'],
    ['MER-503-20GC-P', '工业相机 500万像素 彩色 GigE接口 2/3" CMOS', '大恒图像', 1, '台', 3200, '熊振'],
    ['M1614-MP2', '定焦镜头 16mm F1.4 C口 2/3"', 'Computar', 3, '个', 850, '熊振'],
    ['M2514-MP2', '定焦镜头 25mm F1.4 C口 2/3"', 'Computar', 1, '个', 980, '熊振'],
    ['RL12090-W', '环形光源 120mm外径 90°照射角 白光', 'OPT奥普特', 3, '个', 680, '熊振'],
    ['BL18018-W', '条形光源 180×18mm 白光 高亮', 'OPT奥普特', 2, '个', 850, '熊振'],
    ['DP1024-4CH', '数字光源控制器 4通道 24V/1A', 'OPT奥普特', 2, '台', 1200, '熊振'],
    ['IPC-610L-I7', '视觉工控机 研华IPC-610L i7-12700/32G/512G+2T', '研华', 1, '台', 6800, '熊振'],
    ['U2422H', '显示器 23.8英寸 IPS 1920×1080', '戴尔', 1, '台', 1200, '熊振'],
    ['CA100-AL', '相机安装支架 铝合金 可调角度 兼容M4/M6', '大恒图像', 4, '个', 180, '熊振'],
    ['I350-T4', '千兆网卡 PCIe x4 4口 Intel I350芯片', 'Intel', 1, '个', 580, '熊振'],
    ['NW102-SFTP', '超六类屏蔽网线 2米 蓝色', '绿联', 6, '根', 18, '熊振'],
    ['CG-100-100', '棋盘格标定板 100×100mm 方格6×6 陶瓷基板', 'OPT奥普特', 1, '块', 350, '熊振'],
    ['BP660-25.4', '带通滤光片 BP660nm Φ25.4mm 硬膜', 'Edmund Optics', 3, '片', 420, '熊振'],
    ['PRA-25.4-BK7', '直角棱镜 BK7 25.4mm AR镀膜 400-700nm', 'Thorlabs', 2, '个', 680, '熊振'],
]

start_row3 = 5
for i, row in enumerate(vis_parts):
    r = start_row3 + i
    model, name, brand, qty, unit, price, submitter = row
    ws3.cell(row=r, column=1, value=i + 1)             # 序号
    ws3.cell(row=r, column=2, value=model)               # 型号
    ws3.cell(row=r, column=3, value=name)                # 名称规格
    ws3.cell(row=r, column=4, value=brand)               # 品牌
    ws3.cell(row=r, column=5, value=qty)                 # 单台数量
    ws3.cell(row=r, column=6, value=n_val)               # 装机台数 n
    ws3.cell(row=r, column=7, value=f'=E{r}*F{r}')     # n台数量
    ws3.cell(row=r, column=8, value=unit)                # 单位
    ws3.cell(row=r, column=9, value=price)               # 单价
    ws3.cell(row=r, column=10, value=f'=I{r}*G{r}')    # 总价
    ws3.cell(row=r, column=11, value=submitter)          # 提交人
    ws3.cell(row=r, column=12, value='王工')             # 验收人

print(f'视觉BOM: 写入 {len(vis_parts)} 行')

# Save
wb.save(path)
print(f'\n文件已保存: {path}')
print(f'总计: 外购件{len(parts)}项 + 外加工{len(fab_parts)}项 + 视觉{len(vis_parts)}项 = {len(parts)+len(fab_parts)+len(vis_parts)}项')
