"""向「三工位-15台测试.xlsx」填入电气BOM测试数据"""
import openpyxl
import shutil, os

src_path = r'C:\Users\20210817\Desktop\三工位-15台测试.xlsx'
# 先复制一份再操作，避免锁文件问题
dst_path = r'C:\Users\20210817\Desktop\三工位-15台测试_已填.xlsx'

wb = openpyxl.load_workbook(src_path)
ws = wb['电气BOM']
n_val = 1

elec_parts = [
    # === PLC 控制器 ===
    ['FX5U-80MT/ES', 'PLC主机 FX5U-80MT/ES 40入40出 晶体管', '三菱', 1, '台', 3200, '熊振'],
    ['FX5-16EX/ES', 'PLC输入扩展模块 FX5-16EX/ES 16点DC24V', '三菱', 1, '台', 980, '熊振'],
    ['FX5-16EYR/ES', 'PLC输出扩展模块 FX5-16EYR/ES 16点继电器', '三菱', 2, '台', 850, '熊振'],
    ['FX5-4AD', '模拟量输入模块 FX5-4AD 4通道 电压/电流', '三菱', 1, '台', 1450, '熊振'],
    ['FX5-4DA', '模拟量输出模块 FX5-4DA 4通道', '三菱', 1, '台', 1380, '熊振'],
    ['GT2710-STBA', '触摸屏 GOT2000 GT2710 10.4寸 65536色', '三菱', 1, '台', 4200, '熊振'],
    # === 电源/配电 ===
    ['S-350-24', '开关电源 S-350-24 输入AC220V 输出DC24V 14.6A', '明纬', 2, '台', 185, '熊振'],
    ['S-150-12', '开关电源 S-150-12 输入AC220V 输出DC12V 12.5A', '明纬', 1, '台', 95, '熊振'],
    ['MDR-60-24', '导轨电源 MDR-60-24 DC24V 2.5A DIN安装', '明纬', 2, '台', 128, '熊振'],
    ['NXB-63-C32-2P', '小型断路器 NXB-63 C32 2P 6kA', '正泰', 2, '个', 28, '熊振'],
    ['NXB-63-C16-1P', '小型断路器 NXB-63 C16 1P 6kA', '正泰', 4, '个', 15, '熊振'],
    ['NXB-63-C10-1P', '小型断路器 NXB-63 C10 1P 6kA', '正泰', 3, '个', 13, '熊振'],
    ['NXB-63-C6-2P', '小型断路器 NXB-63 C6 2P 6kA', '正泰', 2, '个', 32, '熊振'],
    ['NXBLE-63-C32-2P', '漏电保护器 NXBLE-63 C32 2P 30mA', '正泰', 1, '个', 55, '熊振'],
    # === 接触器/继电器 ===
    ['CJX2-2510-220V', '交流接触器 CJX2-2510 线圈AC220V 25A', '正泰', 3, '个', 42, '熊振'],
    ['NR2-25-7-10A', '热继电器 NR2-25 整定范围7-10A', '正泰', 3, '个', 28, '熊振'],
    ['JZX-22F/4Z-DC24V', '中间继电器 JZX-22F/4Z 4组转换 DC24V', '正泰', 8, '个', 12, '熊振'],
    ['PYF08A-E', '继电器插座 PYF08A-E 8脚 DIN导轨安装', '正泰', 8, '个', 5, '熊振'],
    ['JQX-13F-2Z-DC24V', '大功率继电器 JQX-13F 2组转换 10A DC24V', '正泰', 4, '个', 15, '熊振'],
    # === 端子排 ===
    ['UK2.5B-GRAY', '接线端子 UK2.5B 灰色 2.5mm² 螺丝压接', '菲尼克斯', 80, '个', 1.5, '熊振'],
    ['UK2.5B-BLUE', '接线端子 UK2.5B 蓝色 N线专用', '菲尼克斯', 40, '个', 1.5, '熊振'],
    ['UK2.5B-PE', '接地端子 UK2.5B-PE 黄绿色', '菲尼克斯', 30, '个', 2.0, '熊振'],
    ['D-UK2.5', '端子端板 D-UK2.5 灰色', '菲尼克斯', 20, '个', 1.2, '熊振'],
    ['ZB6-LGS-1-10', '端子标记条 ZB6 LGS:1-10 白色', '菲尼克斯', 15, '条', 3.5, '熊振'],
    ['UKH50', '大电流端子 UKH50 50mm² 灰色', '菲尼克斯', 6, '个', 8.5, '熊振'],
    # === 传感器 ===
    ['E3Z-D61', '光电传感器 E3Z-D61 漫反射型 NPN 300mm', '欧姆龙', 3, '个', 85, '熊振'],
    ['E3X-NA11-2M', '光纤放大器 E3X-NA11 NPN输出 导线2m', '欧姆龙', 3, '个', 180, '熊振'],
    ['E32-DC200', '光纤探头 E32-DC200 同轴型 M6', '欧姆龙', 3, '个', 120, '熊振'],
    ['TL-Q5MC1', '接近开关 TL-Q5MC1 方形 NPN常开 检测5mm', '欧姆龙', 6, '个', 35, '熊振'],
    ['E2E-X5MF1', '接近开关 E2E-X5MF1 M12 屏蔽型 NPN 5mm', '欧姆龙', 4, '个', 48, '熊振'],
    ['D-M9B-CS1-J', '磁性开关 D-M9B 固态无触点 导线1m', 'SMC', 6, '个', 55, '熊振'],
    ['WLCA12-2N', '限位开关 WLCA12-2N 滚轮摆杆型', '欧姆龙', 6, '个', 42, '熊振'],
    # === 电磁阀 ===
    ['SY5120-5LZD-C6', '电磁阀 SY5120 5通 单线圈 DC24V C6快插', 'SMC', 3, '个', 280, '熊振'],
    ['SY5220-5LZD-C6', '电磁阀 SY5220 5通 双线圈 DC24V C6快插', 'SMC', 2, '个', 310, '熊振'],
    ['SS5Y3-20-02-4ST', '汇流板 SS5Y3-20 阀岛底座 4联 C6快插', 'SMC', 1, '个', 380, '熊振'],
    ['AN103-KM5', '消音器 AN103 快插型 M5螺纹 铜烧结', 'SMC', 4, '个', 8, '熊振'],
    ['AN203-02', '消音器 AN203-02 R1/4螺纹 铜烧结', 'SMC', 2, '个', 12, '熊振'],
    # === 按钮/指示灯 ===
    ['XB2BS542C', '急停按钮 XB2BS542C 红色 旋转复位 1NC', '施耐德', 2, '个', 45, '熊振'],
    ['XB2BA31C', '平头按钮 XB2BA31C 绿色 1NO 复位型', '施耐德', 3, '个', 12, '熊振'],
    ['XB2BA42C', '平头按钮 XB2BA42C 红色 1NC 复位型', '施耐德', 3, '个', 12, '熊振'],
    ['XB2BA51C', '平头按钮 XB2BA51C 黄色 1NO 复位型', '施耐德', 2, '个', 12, '熊振'],
    ['XB2BVM3LC', '指示灯 XB2BVM3LC 绿色 AC220V LED', '施耐德', 3, '个', 18, '熊振'],
    ['XB2BVM4LC', '指示灯 XB2BVM4LC 红色 AC220V LED', '施耐德', 3, '个', 18, '熊振'],
    ['XB2BVM5LC', '指示灯 XB2BVM5LC 黄色 AC220V LED', '施耐德', 1, '个', 18, '熊振'],
    ['XB2BSB4LC', '蜂鸣器 XB2BSB4LC 红色 90dB AC220V', '施耐德', 1, '个', 55, '熊振'],
    ['XB2BD21C', '选择开关 XB2BD21C 2位 自锁 黑色', '施耐德', 2, '个', 22, '熊振'],
    ['XB2BD33C', '选择开关 XB2BD33C 3位 自复位 黑色', '施耐德', 2, '个', 25, '熊振'],
    # === 伺服/步进 ===
    ['SV630PS2R8I', '伺服驱动器 SV630PS2R8I 750W 单相220V', '汇川技术', 1, '台', 1250, '熊振'],
    ['MS1H1-75B30CB-A334Z', '伺服电机 MS1H1 750W 3000rpm 带抱闸', '汇川技术', 1, '台', 980, '熊振'],
    ['S6-L-P021-3.0', '伺服动力线 S6-L-P021 3米', '汇川技术', 1, '根', 120, '熊振'],
    ['S6-L-F021-3.0', '伺服编码器线 S6-L-F021 3米', '汇川技术', 1, '根', 135, '熊振'],
    ['DM542', '步进驱动器 DM542 4.2A 最高128细分', '雷赛智能', 3, '台', 185, '熊振'],
    ['57HS22-A', '步进电机 57HS22 2.2N·m 4A 1.8°', '雷赛智能', 3, '台', 160, '熊振'],
    # === 电气柜/附件 ===
    ['AE-1038.500', '电气柜 AE 1800×800×500mm 前单门 IP54', '威图', 1, '台', 2800, '熊振'],
    ['SZ-2485.200', '安装板 SZ 1750×750mm 镀锌板 2mm', '威图', 1, '块', 320, '熊振'],
    ['VD-80-60-GRAY', '线槽 VD 80×60mm PVC 灰色 2m/根', '凯士士', 15, '根', 18, '熊振'],
    ['VD-60-40-GRAY', '线槽 VD 60×40mm PVC 灰色 2m/根', '凯士士', 10, '根', 12, '熊振'],
    ['TS35-7.5-1000', 'DIN导轨 TS35 7.5mm高 1m/根 镀锌', '国产', 8, '根', 8, '熊振'],
    ['PG-M20×1.5', '电缆格兰头 M20×1.5 尼龙PA66 黑色', '国产', 20, '个', 1.5, '熊振'],
    ['PG-M25×1.5', '电缆格兰头 M25×1.5 尼龙PA66 黑色', '国产', 10, '个', 2.0, '熊振'],
    ['LED-TL-10W', '柜内照明灯 LED 10W AC220V 带门控开关', '施耐德', 1, '个', 85, '熊振'],
    ['KA1238HA2-220V', '散热风扇 120×120×38mm AC220V 滚珠轴承', '卡固', 2, '个', 55, '熊振'],
    ['KF-120', '风扇过滤网 120×120mm 配KA1238', '卡固', 2, '个', 12, '熊振'],
    # === 线缆 ===
    ['RVV-4×2.5', '动力电缆 RVV 4芯×2.5mm² 黑色 450/750V', '起帆', 50, '米', 8.5, '熊振'],
    ['RVV-10×0.5', '控制电缆 RVV 10芯×0.5mm² 黑色', '起帆', 30, '米', 6.5, '熊振'],
    ['RVVP-4×0.5', '屏蔽电缆 RVVP 4芯×0.5mm² 铜网屏蔽', '起帆', 40, '米', 5.5, '熊振'],
    ['CAT5E-SFTP-305M', '超五类屏蔽网线 SFTP 4对 305米/箱', '安普AMP', 1, '箱', 580, '熊振'],
    ['RJ45-CAT5E', '水晶头 RJ45 超五类 屏蔽 100个/盒', '安普AMP', 1, '盒', 45, '熊振'],
    ['BVR-2.5-RED', '单芯软线 BVR 2.5mm² 红色 100米/卷', '起帆', 2, '卷', 95, '熊振'],
    ['BVR-2.5-BLUE', '单芯软线 BVR 2.5mm² 蓝色 100米/卷', '起帆', 2, '卷', 95, '熊振'],
    ['BVR-1.5-BLACK', '单芯软线 BVR 1.5mm² 黑色 100米/卷', '起帆', 3, '卷', 58, '熊振'],
]

start_row = 5
for i, row in enumerate(elec_parts):
    r = start_row + i
    model, name, brand, qty, unit, price, submitter = row
    ws.cell(row=r, column=1, value=i + 1)
    ws.cell(row=r, column=2, value=model)
    ws.cell(row=r, column=3, value=name)
    ws.cell(row=r, column=4, value=brand)
    ws.cell(row=r, column=5, value=qty)
    ws.cell(row=r, column=6, value=n_val)
    ws.cell(row=r, column=7, value=f'=E{r}*F{r}')
    ws.cell(row=r, column=8, value=unit)
    ws.cell(row=r, column=9, value=price)
    ws.cell(row=r, column=10, value=f'=I{r}*G{r}')
    ws.cell(row=r, column=11, value=submitter)
    ws.cell(row=r, column=12, value='王工')

wb.save(dst_path)

# Count categories
cats = {}
for row in elec_parts:
    name = row[2]
    if 'PLC' in name: cat = 'PLC控制器'
    elif '触摸' in name: cat = '触摸屏HMI'
    elif '电源' in name: cat = '电源'
    elif '断路' in name or '漏电' in name: cat = '断路器/漏保'
    elif '接触' in name or '继电器' in name or '插座' in name: cat = '接触器继电器'
    elif '端子' in name or '标记' in name: cat = '端子排'
    elif '接地' in name: cat = '接地端子'
    elif '传感' in name or '接近' in name or '磁性' in name or '限位' in name: cat = '传感器'
    elif '光纤' in name or '探头' in name: cat = '光纤附件'
    elif '阀' in name or '汇流' in name: cat = '电磁阀'
    elif '消音' in name: cat = '消音器'
    elif '按钮' in name or '指示' in name or '蜂鸣' in name or '选择' in name: cat = '按钮指示灯'
    elif '伺服' in name: cat = '伺服驱动'
    elif '步进' in name: cat = '步进驱动'
    elif '柜' in name or '安装板' in name or '线槽' in name or '导轨' in name or '格兰' in name or '照明' in name or '风扇' in name or '过滤' in name: cat = '电气柜附件'
    else: cat = '线缆'
    cats[cat] = cats.get(cat, 0) + 1

print(f'电气BOM: 写入 {len(elec_parts)} 行 (Row {start_row}-{start_row+len(elec_parts)-1})')
for cat, cnt in sorted(cats.items()):
    print(f'  {cat}: {cnt}项')

total = sum(row[3]*row[5] for row in elec_parts)
print(f'\n总金额(单价×数量): ¥{total:,.0f}')
print(f'\n文件已保存: {dst_path}')
