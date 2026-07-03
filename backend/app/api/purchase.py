"""采购管理 API"""
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from typing import Optional
import os, uuid

from app.core.database import get_db
from app.core.config import UPLOAD_DIR
from app.models.order import PurchaseOrder
from app.models.supplier import Supplier
from app.models.material import MaterialMaster
from app.models.bom import BomHeader, BomLine
from app.models.inventory import InventoryRecord, InventoryTransaction, Warehouse

router = APIRouter(prefix="/api/purchase", tags=["采购管理"])


@router.get("/orders/all")
def list_purchase_orders_all(db: Session = Depends(get_db)):
    """获取全部采购订单（不分页，供折叠视图使用）"""
    orders = db.query(PurchaseOrder).order_by(PurchaseOrder.id.desc()).all()
    return {
        "items": [
            {
                "id": o.id, "po_number": o.po_number,
                "supplier_id": o.supplier_id,
                "supplier_name": o.supplier.supplier_name if o.supplier else "",
                "item_id": o.item_id,
                "material_code": o.item.material_code if o.item else "",
                "material_name": o.item.material_name if o.item else "",
                "order_qty": o.order_qty, "received_qty": o.received_qty,
                "unit_price": o.unit_price or 0, "total_amount": o.total_amount or 0,
                "brand": o.brand or "", "submitter": o.submitter or "",
                "supplier_link": o.supplier_link or "",
                "due_date": o.due_date.isoformat() if o.due_date else None,
                "status": o.status, "source_type": o.source_type,
                "priority": o.priority,
                "created_at": o.created_at.isoformat() if o.created_at else None,
            }
            for o in orders
        ],
        "total": len(orders),
    }


@router.get("/orders")
def list_purchase_orders(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=1000),
    status: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """采购订单列表"""
    query = db.query(PurchaseOrder)

    if status:
        query = query.filter(PurchaseOrder.status == status)
    if keyword:
        query = query.join(MaterialMaster, PurchaseOrder.item_id == MaterialMaster.id).filter(
            (PurchaseOrder.po_number.ilike(f"%{keyword}%")) |
            (MaterialMaster.material_code.ilike(f"%{keyword}%")) |
            (MaterialMaster.material_name.ilike(f"%{keyword}%"))
        )

    total = query.count()
    orders = query.order_by(PurchaseOrder.created_at.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    return {
        "items": [
            {
                "id": o.id, "po_number": o.po_number,
                "supplier_id": o.supplier_id,
                "supplier_name": o.supplier.supplier_name if o.supplier else "",
                "item_id": o.item_id,
                "material_code": o.item.material_code if o.item else "",
                "material_name": o.item.material_name if o.item else "",
                "unit": o.item.unit if o.item else "",
                "order_qty": o.order_qty, "received_qty": o.received_qty,
                "unit_price": o.unit_price or 0, "total_amount": o.total_amount or 0,
                "brand": o.brand or "", "submitter": o.submitter or "",
                "supplier_link": o.supplier_link or "",
                "due_date": o.due_date.isoformat() if o.due_date else None,
                "status": o.status, "source_type": o.source_type,
                "priority": o.priority, "remark": o.remark,
                "created_at": o.created_at.isoformat() if o.created_at else None,
            }
            for o in orders
        ],
        "total": total, "page": page, "page_size": page_size,
    }


@router.post("/orders")
def create_purchase_order(data: dict, db: Session = Depends(get_db)):
    """创建采购订单"""
    # 自动生成唯一单号
    if "po_number" not in data or not data.get("po_number"):
        today_str = date.today().strftime("%Y%m%d")
        last_po = db.query(PurchaseOrder).filter(
            PurchaseOrder.po_number.like(f"PO-{today_str}-%")
        ).order_by(PurchaseOrder.po_number.desc()).first()
        if last_po:
            last_seq = int(last_po.po_number.split("-")[-1])
            seq = last_seq + 1
        else:
            seq = 1
        po_number = f"PO-{today_str}-{seq:04d}"
    else:
        po_number = data["po_number"]

    po = PurchaseOrder(
        po_number=po_number,
        supplier_id=data["supplier_id"],
        item_id=data["item_id"],
        order_qty=data["order_qty"],
        due_date=date.fromisoformat(data["due_date"]) if isinstance(data["due_date"], str) else data["due_date"],
        status=data.get("status", "申请"),
        source_type=data.get("source_type", "手动"),
        priority=data.get("priority", 0),
        remark=data.get("remark", ""),
    )
    db.add(po)
    db.commit()
    return {"success": True, "data": {"id": po.id, "po_number": po.po_number}}


@router.put("/orders/{order_id}/status")
def update_po_status(order_id: int, data: dict, db: Session = Depends(get_db)):
    """更新采购单状态 — 到货时自动同步库存"""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="采购单不存在")

    old_status = po.status
    old_received = po.received_qty or 0
    new_status = data["status"]
    new_received = data.get("received_qty", old_received)
    if new_received is None:
        new_received = old_received
    
    # 计算本次到货增量
    delta = new_received - old_received

    po.status = new_status
    po.received_qty = new_received
    
    # 自动入库：当状态变为部分到货或全部到货且数量有增加
    if new_status in ("部分到货", "全部到货") and delta > 0:
        # 找到或创建仓库
        wh = db.query(Warehouse).filter(Warehouse.warehouse_code == "WH01").first()
        if not wh:
            wh = Warehouse(warehouse_code="WH01", warehouse_name="主仓库")
            db.add(wh)
            db.flush()
        
        # 更新库存记录
        inv = db.query(InventoryRecord).filter(
            InventoryRecord.item_id == po.item_id,
            InventoryRecord.warehouse_id == wh.id,
        ).first()
        if not inv:
            inv = InventoryRecord(item_id=po.item_id, warehouse_id=wh.id)
            db.add(inv)
            db.flush()
        
        inv.on_hand_qty += delta
        
        # 记录入库流水
        tx = InventoryTransaction(
            item_id=po.item_id,
            warehouse_id=wh.id,
            transaction_type="采购入库",
            quantity=delta,
            reference_no=po.po_number,
            operator=data.get("operator", ""),
            remark=f"采购单 {po.po_number} 到货, 供应商:{po.supplier.supplier_name if po.supplier else ''}",
        )
        db.add(tx)

    db.commit()
    
    msg = f"状态更新为: {po.status}"
    if delta > 0:
        msg = f"收货 {delta} 件, 累计 {new_received}/{po.order_qty}, 状态: {po.status}, 已自动入库"
    
    return {"success": True, "message": msg, "data": {"delta": delta, "total_received": new_received}}


@router.delete("/orders/{order_id}")
def delete_purchase_order(order_id: int, db: Session = Depends(get_db)):
    """删除采购单"""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if po and po.status in ("申请", "已下单"):
        db.delete(po)
        db.commit()
        return {"success": True}
    raise HTTPException(status_code=400, detail="只能删除'已下单'状态的采购单")


@router.post("/orders/import/excel")
async def import_purchase_orders(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Excel批量导入采购订单"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    import openpyxl
    file_path = os.path.join(UPLOAD_DIR, f"po_import_{uuid.uuid4().hex}.xlsx")
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    expected = ["物料编码", "数量", "供应商编码", "交货日期", "备注"]
    col_map = {h: i for i, h in enumerate(header_row)}

    missing = [h for h in expected if h not in col_map]
    if missing:
        os.remove(file_path)
        raise HTTPException(status_code=400, detail=f"缺少列: {', '.join(missing)}")

    imported = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            material_code = str(row[col_map["物料编码"]]).strip()
            qty = float(row[col_map["数量"]])
            supplier_code = str(row[col_map["供应商编码"]]).strip()
            due = str(row[col_map["交货日期"]]).strip()
            remark = str(row[col_map.get("备注", 4)]).strip() if col_map.get("备注", 4) < len(row) else ""

            material = db.query(MaterialMaster).filter(MaterialMaster.material_code == material_code).first()
            if not material:
                errors.append(f"第{i}行: 物料编码 {material_code} 不存在")
                continue
            supplier = db.query(Supplier).filter(Supplier.supplier_code == supplier_code).first()
            if not supplier:
                errors.append(f"第{i}行: 供应商编码 {supplier_code} 不存在")
                continue

            # 自动生成单号
            today_str = date.today().strftime("%Y%m%d")
            last_po = db.query(PurchaseOrder).filter(
                PurchaseOrder.po_number.like(f"PO-{today_str}-%")
            ).order_by(PurchaseOrder.po_number.desc()).first()
            seq = int(last_po.po_number.split("-")[-1]) + 1 if last_po else 1
            po_number = f"PO-{today_str}-{seq:04d}"

            po = PurchaseOrder(
                po_number=po_number,
                supplier_id=supplier.id,
                item_id=material.id,
                order_qty=qty,
                due_date=date.fromisoformat(due),
                status="申请",
                source_type="导入",
                remark=remark,
            )
            db.add(po)
            imported += 1
        except Exception as e:
            errors.append(f"第{i}行: {str(e)}")

    if imported:
        db.commit()
    os.remove(file_path)
    return {
        "success": True,
        "message": f"成功导入 {imported} 条采购订单",
        "errors": errors[:10],
        "total_errors": len(errors),
    }


@router.get("/orders/import/template")
def download_import_template():
    """下载采购订单导入模板"""
    import openpyxl
    from fastapi.responses import FileResponse

    file_path = os.path.join(UPLOAD_DIR, "po_import_template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采购订单导入"
    ws.append(["物料编码", "数量", "供应商编码", "交货日期", "备注"])
    ws.append(["RM-001", 100, "SUP001", "2026-07-15", "急用"])
    ws.append(["RM-002", 200, "SUP001", "2026-07-20", ""])
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 20
    wb.save(file_path)
    return FileResponse(file_path, filename="采购订单导入模板.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ====== 供应商管理 ======
@router.get("/suppliers")
def list_suppliers(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1),
    keyword: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """供应商列表"""
    query = db.query(Supplier).filter(Supplier.is_active == True)
    if keyword:
        query = query.filter(
            (Supplier.supplier_code.ilike(f"%{keyword}%")) |
            (Supplier.supplier_name.ilike(f"%{keyword}%"))
        )
    total = query.count()
    suppliers = query.offset((page-1)*page_size).limit(page_size).all()

    return {
        "items": [
            {"id": s.id, "supplier_code": s.supplier_code, "supplier_name": s.supplier_name,
             "contact_person": s.contact_person, "contact_phone": s.contact_phone,
             "lead_time_days": s.lead_time_days, "purchase_link": s.purchase_link or ""}
            for s in suppliers
        ],
        "total": total,
    }


@router.post("/suppliers")
def create_supplier(data: dict, db: Session = Depends(get_db)):
    """创建供应商"""
    s = Supplier(**data)
    db.add(s)
    db.commit()
    return {"success": True, "data": {"id": s.id}}


@router.put("/suppliers/{supplier_id}")
def update_supplier(supplier_id: int, data: dict, db: Session = Depends(get_db)):
    """更新供应商"""
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="供应商不存在")
    for k, v in data.items():
        if hasattr(s, k):
            setattr(s, k, v)
    db.commit()
    return {"success": True, "message": "已更新"}


@router.delete("/suppliers/{supplier_id}")
def delete_supplier(supplier_id: int, db: Session = Depends(get_db)):
    """删除供应商（软删除）"""
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="供应商不存在")
    s.is_active = False
    db.commit()
    return {"success": True, "message": "已删除"}


# ====== BOM同步采购 ======

@router.post("/sync-from-bom")
def sync_purchase_from_bom(data: dict, db: Session = Depends(get_db)):
    """
    从BOM同步生成采购申请
    
    对BOM中所有 is_purchased=True 的零件，自动创建采购申请。
    如果供应商不存在，自动创建供应商。
    
    请求体:
    {
        "bom_header_id": 1,     // 或
        "product_id": 1,         // 二选一
        "auto_approve": false    // 是否自动审批
    }
    """
    bom_header_id = data.get("bom_header_id")
    product_id = data.get("product_id")
    auto_approve = data.get("auto_approve", False)
    
    if not bom_header_id and not product_id:
        raise HTTPException(status_code=400, detail="请提供 bom_header_id 或 product_id")
    
    # 找到BOM
    if bom_header_id:
        header = db.query(BomHeader).filter(BomHeader.id == bom_header_id).first()
    else:
        header = db.query(BomHeader).filter(
            BomHeader.product_id == product_id
        ).order_by(BomHeader.id.desc()).first()
    
    if not header:
        raise HTTPException(status_code=404, detail="BOM不存在")
    
    # 收集BOM中的所有采购件
    from app.models.bom import BomLine
    from sqlalchemy import func
    
    created_po = 0
    created_supplier = 0
    skipped = 0
    errors = []
    
    # 递归获取所有子物料
    def get_all_children(parent_id, visited=None):
        if visited is None:
            visited = set()
        items = []
        lines = db.query(BomLine).filter(
            BomLine.bom_header_id == header.id,
            BomLine.parent_item_id == parent_id
        ).all()
        for line in lines:
            if line.item_id not in visited:
                visited.add(line.item_id)
                items.append(line.item)
                items.extend(get_all_children(line.item_id, visited))
            else:
                items.extend(get_all_children(line.item_id, visited))
        return items
    
    # 先获取顶层物料下的直接子物料
    from app.models.material import MaterialMaster as MM
    top_children = db.query(BomLine).filter(
        BomLine.bom_header_id == header.id,
        BomLine.parent_item_id == header.product_id
    ).all()
    
    all_items = []
    visited = set()
    for line in top_children:
        if line.item_id not in visited:
            visited.add(line.item_id)
            item = db.query(MM).filter(MM.id == line.item_id).first()
            if item:
                all_items.append((item, line.quantity))
    
    # 递归展开模块层
    all_parts = []
    seen = set()
    queue = [(item, qty) for item, qty in all_items]
    while queue:
        mat, multiplier = queue.pop(0)
        if mat.id in seen:
            continue
        seen.add(mat.id)
        
        if mat.is_purchased and mat.level_type == "零件":
            all_parts.append((mat, multiplier))
        elif mat.level_type in ("模块", "产品"):
            # 展开子物料
            children = db.query(BomLine).filter(
                BomLine.bom_header_id == header.id,
                BomLine.parent_item_id == mat.id
            ).all()
            for child in children:
                child_mat = db.query(MM).filter(MM.id == child.item_id).first()
                if child_mat:
                    queue.append((child_mat, multiplier * child.quantity))
    
    # 收集品牌列表
    from collections import Counter
    brand_set = set()
    
    # 创建采购申请
    today_str = date.today().strftime("%Y%m%d")
    last_po = db.query(PurchaseOrder).filter(
        PurchaseOrder.po_number.like(f"PO-{today_str}-%")
    ).order_by(PurchaseOrder.po_number.desc()).first()
    po_seq = int(last_po.po_number.split("-")[-1]) + 1 if last_po else 1
    
    # 供应商缓存
    supplier_cache = {}
    for s in db.query(Supplier).all():
        supplier_cache[s.supplier_name] = s
    
    for mat, qty in all_parts:
        # 检查是否已有采购申请
        existing = db.query(PurchaseOrder).filter(
            PurchaseOrder.item_id == mat.id,
            PurchaseOrder.status.in_(["已下单", "部分到货"]),
        ).first()
        if existing:
            skipped += 1
            continue
        
        # 品牌：从 specification 取
        brand = (mat.specification or "").strip()
        
        # 供应商：优先用品牌名创建/匹配
        supp_name = brand if brand else "默认供应商"
        if supp_name in supplier_cache:
            supplier = supplier_cache[supp_name]
        else:
            supp_code = f"SUP-{supp_name[:15]}" if len(supp_name) < 20 else f"SUP-{supp_name[:10]}"
            existing_sup = db.query(Supplier).filter(Supplier.supplier_code == supp_code).first()
            if existing_sup:
                supplier = existing_sup
            else:
                supplier = Supplier(supplier_code=supp_code, supplier_name=supp_name, lead_time_days=5)
                db.add(supplier)
                db.flush()
                created_supplier += 1
            supplier_cache[supp_name] = supplier
        
        # 如果物料有参考链接，更新供应商的采购链接
        ref_link = getattr(mat, 'reference_link', '') or ''
        if ref_link:
            if not supplier.purchase_link:
                supplier.purchase_link = ref_link
            # also store on the PO
        else:
            ref_link = ''
        
        # 创建采购单
        po = PurchaseOrder(
            po_number=f"PO-{today_str}-{po_seq:04d}",
            supplier_id=supplier.id,
            item_id=mat.id,
            order_qty=qty,
            due_date=date.today(),
            status="已下单",
            source_type="BOM同步",
            brand=brand,
            unit_price=getattr(mat, 'reference_unit_price', 0) or 0,
            total_amount=(getattr(mat, 'reference_unit_price', 0) or 0) * qty,
            submitter=getattr(mat, 'reference_submitter', '') or '',
            supplier_link=getattr(mat, 'reference_link', '') or '',
            priority=0,
            remark="",
        )
        db.add(po)
        created_po += 1
        po_seq += 1
    
    db.commit()
    
    return {
        "success": True,
        "message": f"同步完成: 采购申请 {created_po} 笔, 跳过(已存在) {skipped} 笔, 新建供应商 {created_supplier} 个",
        "data": {
            "created_orders": created_po,
            "skipped": skipped,
            "new_suppliers": created_supplier,
            "total_parts_found": len(all_parts),
        },
    }
