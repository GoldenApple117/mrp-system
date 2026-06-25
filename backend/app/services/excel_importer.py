"""Excel 导入服务 — BOM层级导入、物料导入、库存导入"""
import os
from datetime import date
from typing import List, Dict, Tuple, Any
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.bom_exploder import BomExploder


class ExcelImporter:
    """Excel数据导入器"""

    # BOM导入的列名映射
    BOM_COLUMN_MAP = {
        "父物料编码": "parent_code",
        "父物料名称": "parent_name",
        "子物料编码": "child_code",
        "子物料名称": "child_name",
        "子物料规格": "child_spec",
        "用量": "quantity_per",
        "位号": "position",
        "损耗率": "scrap_rate",
        "备注": "remark",
    }

    # 物料导入的列名映射
    MATERIAL_COLUMN_MAP = {
        "物料编码": "material_code",
        "物料名称": "material_name",
        "规格型号": "specification",
        "单位": "unit",
        "物料类型": "material_type",
        "提前期": "lead_time",
        "安全库存": "safety_stock",
        "批量规则": "lot_size_rule",
        "批量数量": "lot_size_qty",
        "最小订货量": "min_order_qty",
        "最大订货量": "max_order_qty",
        "损耗率": "scrap_rate",
        "是否外购": "is_purchased",
    }

    def __init__(self, file_path: str):
        self.file_path = file_path
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

    def read_dataframe(self, sheet_name: str = 0) -> pd.DataFrame:
        """读取Excel为DataFrame"""
        return pd.read_excel(self.file_path, sheet_name=sheet_name)

    def import_materials(self, sheet_name: str = "物料主数据") -> Tuple[List[dict], List[str]]:
        """
        导入物料主数据

        Returns:
            (materials列表, 错误列表)
        """
        df = self.read_dataframe(sheet_name)

        # 标准化列名
        df = df.rename(columns={v: k for k, v in self.MATERIAL_COLUMN_MAP.items()
                                if k in self.MATERIAL_COLUMN_MAP})
        # 反向映射（中文→英文）
        reverse_map = {v: k for k, v in self.MATERIAL_COLUMN_MAP.items()}
        # 实际列名映射
        actual_map = {}
        for col in df.columns:
            if col in reverse_map:
                actual_map[col] = reverse_map[col]

        materials = []
        errors = []

        for idx, row in df.iterrows():
            try:
                mat = {
                    "material_code": str(row.get("物料编码", "")).strip(),
                    "material_name": str(row.get("物料名称", "")).strip(),
                    "specification": str(row.get("规格型号", "")) if pd.notna(row.get("规格型号")) else "",
                    "unit": str(row.get("单位", "个")).strip(),
                    "material_type": str(row.get("物料类型", "原材料")).strip(),
                }

                if not mat["material_code"] or not mat["material_name"]:
                    errors.append(f"第{idx+2}行: 物料编码或物料名称为空")
                    continue

                mat["lead_time"] = int(row.get("提前期", 0)) if pd.notna(row.get("提前期")) else 0
                mat["safety_stock"] = float(row.get("安全库存", 0)) if pd.notna(row.get("安全库存")) else 0
                mat["lot_size_rule"] = str(row.get("批量规则", "LFL")).strip().upper()
                mat["lot_size_qty"] = float(row.get("批量数量", 1)) if pd.notna(row.get("批量数量")) else 1
                mat["min_order_qty"] = float(row.get("最小订货量", 0)) if pd.notna(row.get("最小订货量")) else 0
                mat["max_order_qty"] = float(row.get("最大订货量", 0)) if pd.notna(row.get("最大订货量")) else 0
                mat["scrap_rate"] = float(row.get("损耗率", 0)) if pd.notna(row.get("损耗率")) else 0
                is_purchased_val = str(row.get("是否外购", "是")).strip()
                mat["is_purchased"] = is_purchased_val in ("是", "Y", "y", "1", "True", "true")

                materials.append(mat)
            except Exception as e:
                errors.append(f"第{idx+2}行: 数据解析错误 - {str(e)}")

        return materials, errors

    def import_bom(self, sheet_name: str = "BOM") -> Tuple[List[dict], List[str]]:
        """
        导入层级BOM

        Excel格式（父子结构，每行一对父子关系）：
        父物料编码 | 父物料名称 | 子物料编码 | 子物料名称 | 用量 | 位号 | 备注

        Returns:
            (bom_lines列表, 错误列表)
        """
        df = self.read_dataframe(sheet_name)

        bom_lines = []
        errors = []

        for idx, row in df.iterrows():
            try:
                parent_code = str(row.get("父物料编码", "")).strip()
                parent_name = str(row.get("父物料名称", "")) if pd.notna(row.get("父物料名称")) else ""
                child_code = str(row.get("子物料编码", "")).strip()
                child_name = str(row.get("子物料名称", "")) if pd.notna(row.get("子物料名称")) else ""
                child_spec = str(row.get("子物料规格", "")) if pd.notna(row.get("子物料规格")) else ""

                if not child_code:
                    errors.append(f"第{idx+2}行: 子物料编码为空，跳过")
                    continue

                qty_per = float(row.get("用量", 1)) if pd.notna(row.get("用量")) else 1
                position = str(row.get("位号", "")) if pd.notna(row.get("位号")) else ""

                remark = str(row.get("备注", "")) if pd.notna(row.get("备注")) else ""
                is_substitute = "替代" in remark or "替换" in remark

                scrap_rate = float(row.get("损耗率", 0)) if pd.notna(row.get("损耗率", 0)) else 0

                bom_lines.append({
                    "parent_code": parent_code,
                    "parent_name": str(parent_name).strip(),
                    "child_code": child_code,
                    "child_name": str(child_name).strip(),
                    "child_spec": str(child_spec).strip(),
                    "quantity_per": qty_per,
                    "position": str(position).strip(),
                    "scrap_rate": scrap_rate,
                    "is_substitute": is_substitute,
                    "remark": str(remark).strip(),
                })
            except Exception as e:
                errors.append(f"第{idx+2}行: 数据解析错误 - {str(e)}")

        # 验证循环引用
        if bom_lines and not errors:
            exploder = BomExploder(bom_lines)
            cycle = exploder.detect_cycle()
            if cycle:
                cycle_str = " → ".join(cycle)
                errors.append(f"发现循环引用: {cycle_str}")

        return bom_lines, errors

    @staticmethod
    def generate_bom_template(output_path: str) -> str:
        """生成BOM导入模板文件"""
        from openpyxl import Workbook

        wb = Workbook()

        # Sheet 1: BOM 模板
        ws_bom = wb.active
        ws_bom.title = "BOM"

        headers = ["父物料编码", "父物料名称", "子物料编码", "子物料名称",
                    "子物料规格", "用量", "位号", "损耗率(%)", "备注"]
        for col, h in enumerate(headers, 1):
            ws_bom.cell(row=1, column=col, value=h)

        # 示例数据
        example = [
            ["FG-001", "成品A", "SA-001", "部件A1", "规格A1", 1, "A1", 0, ""],
            ["FG-001", "成品A", "SA-002", "部件A2", "规格A2", 2, "A2", 0, ""],
            ["SA-001", "部件A1", "RM-001", "原材料X", "规格X", 3, "", 0, ""],
            ["SA-001", "部件A1", "RM-002", "原材料Y", "规格Y", 1, "", 0, ""],
            ["SA-002", "部件A2", "RM-001", "原材料X", "规格X", 2, "", 0, "替代料"],
            ["SA-002", "部件A2", "RM-003", "原材料Z", "规格Z", 4, "", 0, ""],
        ]
        for r, row_data in enumerate(example, 2):
            for c, val in enumerate(row_data, 1):
                ws_bom.cell(row=r, column=c, value=val)

        # Sheet 2: 物料主数据
        ws_mat = wb.create_sheet("物料主数据")
        mat_headers = ["物料编码", "物料名称", "规格型号", "单位", "物料类型",
                       "提前期", "安全库存", "批量规则", "批量数量",
                       "最小订货量", "最大订货量", "损耗率", "是否外购"]
        for col, h in enumerate(mat_headers, 1):
            ws_mat.cell(row=1, column=col, value=h)

        mat_examples = [
            ["FG-001", "成品A", "", "台", "成品", 5, 10, "LFL", 1, 0, 0, 0, "否"],
            ["SA-001", "部件A1", "", "个", "半成品", 3, 20, "FOQ", 50, 0, 0, 2, "否"],
            ["SA-002", "部件A2", "", "个", "半成品", 3, 20, "FOQ", 50, 0, 0, 2, "否"],
            ["RM-001", "原材料X", "", "kg", "原材料", 2, 50, "EOQ", 200, 0, 0, 0, "是"],
            ["RM-002", "原材料Y", "", "个", "原材料", 1, 30, "LFL", 1, 0, 0, 0, "是"],
            ["RM-003", "原材料Z", "", "m", "原材料", 3, 40, "MULT", 100, 0, 0, 0, "是"],
        ]
        for r, row_data in enumerate(mat_examples, 2):
            for c, val in enumerate(row_data, 1):
                ws_mat.cell(row=r, column=c, value=val)

        # 设置列宽
        for ws in [ws_bom, ws_mat]:
            for col in range(1, 20):
                ws.column_dimensions[get_column_letter(col)].width = 14

        wb.save(output_path)
        return output_path
