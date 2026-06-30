"""采购BOM导入器 — 针对固定模板Excel解析
适用模板：三工位-15台.xlsx 格式
结构：产品→模块→零件 三层级
"""
import os, re, uuid
from datetime import date, datetime
from collections import OrderedDict
import openpyxl
from sqlalchemy.orm import Session

from app.models.material import MaterialMaster
from app.models.bom import BomHeader, BomLine


# 工作表名→模块名映射规则（关键词匹配）
SHEET_MAP = [
    (r"外购", "外购件模块"),
    (r"加工|机加|钣金", "机加件模块"),
    (r"电气", "电气模块"),
    (r"视觉", "视觉模块"),
    (r"量具|工具", "量具模块"),
]

# 要跳过的工作表
SKIP_PATTERNS = [r"费用", r"售后", r"Sheet", r"维护", r"合计"]

# 机加件表使用不同的列映射（图号代替型号）
SHEET_SPECIAL_COLUMNS = {
    "机加件模块": {"model_col": 2, "name_col": 3},  # 图号列、零件名称列
}


def _identify_module(sheet_name: str) -> str | None:
    """通过工作表名识别对应的模块名"""
    for pattern, module_name in SHEET_MAP:
        if re.search(pattern, sheet_name):
            return module_name
    return None


def _should_skip(sheet_name: str) -> bool:
    """判断是否应该跳过该工作表"""
    for pattern in SKIP_PATTERNS:
        if re.search(pattern, sheet_name):
            return True
    return False


def _extract_product_name(file_name: str) -> str:
    """从文件名提取产品名（去掉扩展名）"""
    name = os.path.splitext(os.path.basename(file_name))[0]
    return name.strip()


def _is_real_data_row(ws, row: int, model_col: int, name_col: int) -> bool:
    """判断是否为有效数据行（跳过公式行、空行、汇总行）"""
    seq = ws.cell(row, 1).value
    model = ws.cell(row, model_col).value
    name = ws.cell(row, name_col).value

    # 跳过公式行
    if isinstance(seq, str) and seq.startswith("="):
        return False
    # 跳过无名称行
    if not name or not isinstance(name, str) or not name.strip():
        return False
    # 跳过汇总行
    if "合计" in name or "总计" in name:
        return False
    return True


def _generate_code(db: Session, prefix: str = "MAT") -> str:
    """自动生成物料编码"""
    today_str = datetime.now().strftime("%Y%m%d")
    code_prefix = f"{prefix}-{today_str}-"
    last = db.query(MaterialMaster).filter(
        MaterialMaster.material_code.like(f"{code_prefix}%")
    ).order_by(MaterialMaster.material_code.desc()).first()
    seq = int(last.material_code.split("-")[-1]) + 1 if last else 1
    return f"{code_prefix}{seq:05d}"


def import_procurement_bom(file_path: str, db: Session) -> dict:
    """
    主入口：导入采购BOM Excel
    返回值: {"success": bool, "message": str, "stats": {...}, "errors": []}
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    product_name = _extract_product_name(file_path)

    stats = {"product": 0, "modules": 0, "parts": 0, "bom_lines": 0, "skipped": 0}
    errors = []

    # Step 1: 扫描工作表，识别模块和数据
    sheets_info = []  # [(sheet_name, module_name, data_rows)]
    for ws_name in wb.sheetnames:
        if _should_skip(ws_name):
            stats["skipped"] += 1
            continue
        module_name = _identify_module(ws_name)
        if not module_name:
            stats["skipped"] += 1
            continue
        sheets_info.append((ws_name, module_name))

    if not sheets_info:
        return {"success": False, "message": "未找到符合模板的工作表", "stats": stats, "errors": ["文件中没有可导入的BOM数据"]}

    # Step 2: 创建或查找产品
    product = db.query(MaterialMaster).filter(
        MaterialMaster.material_code == f"PROD-{product_name}",
        MaterialMaster.level_type == "产品",
    ).first()
    if not product:
        product = MaterialMaster(
            material_code=f"PROD-{product_name}",
            material_name=product_name,
            unit="台",
            material_type="成品",
            level_type="产品",
            lead_time=0,
            safety_stock=0,
            lot_size_rule="LFL",
            is_purchased=False,
        )
        db.add(product)
        db.flush()
        stats["product"] = 1

    # Step 3: 遍历每个工作表，创建模块和零件
    module_map = {}  # module_name -> MaterialMaster
    created_codes = set()  # 用于行内去重

    for ws_name, module_name in sheets_info:
        ws = wb[ws_name]

        # 创建或查找模块
        mod = db.query(MaterialMaster).filter(
            MaterialMaster.material_code == f"MOD-{module_name}",
            MaterialMaster.level_type == "模块",
        ).first()
        if not mod:
            mod = MaterialMaster(
                material_code=f"MOD-{module_name}",
                material_name=module_name,
                unit="个",
                material_type="模块",
                level_type="模块",
                lead_time=0,
                safety_stock=0,
                lot_size_rule="LFL",
                is_purchased=False,
            )
            db.add(mod)
            db.flush()
            stats["modules"] += 1
        module_map[module_name] = mod

        # 确定列映射（机加件表特殊处理）
        is_special = module_name in SHEET_SPECIAL_COLUMNS
        model_col = SHEET_SPECIAL_COLUMNS[module_name]["model_col"] if is_special else 2  # B列
        name_col = SHEET_SPECIAL_COLUMNS[module_name]["name_col"] if is_special else 3   # C列
        qty_col = 5  # E列

        # 找到表头行（含"序号"的行）
        header_row = None
        for r in range(1, 10):
            if ws.cell(r, 1).value and str(ws.cell(r, 1).value).strip() == "序号":
                header_row = r
                break
        if not header_row:
            errors.append(f"{ws_name}: 未找到表头行，跳过")
            continue

        ws._current_row = header_row + 2  # 数据从表头下2行开始

        # 遍历数据行
        for r in range(header_row + 2, ws.max_row + 1):
            if not _is_real_data_row(ws, r, model_col, name_col):
                continue

            model_val = ws.cell(r, model_col).value
            name_val = ws.cell(r, name_col).value
            qty_val = ws.cell(r, qty_col).value

            part_name = str(name_val).strip() if name_val else ""
            if not part_name:
                continue

            # 确定物料编码：优先用型号列
            part_code = str(model_val).strip() if model_val and str(model_val).strip() else ""
            if not part_code:
                part_code = _generate_code(db)
                part_name_with_code = part_name[:50]
            else:
                part_name_with_code = part_code[:50]
                # 如果编码过长或含有特殊字符，适当处理
                part_code = re.sub(r'[\\/:*?"<>|]', '_', part_code)[:50]

            # 去重：同模块下同一编码只创建一次
            dedup_key = f"{module_name}:{part_code}"
            if dedup_key in created_codes:
                continue
            created_codes.add(dedup_key)

            # 创建零件
            part = db.query(MaterialMaster).filter(
                MaterialMaster.material_code == part_code,
            ).first()
            if not part:
                part = MaterialMaster(
                    material_code=part_code,
                    material_name=part_name[:200],
                    specification=part_name[:500] if part_name != part_code else "",
                    unit="个",
                    material_type="原材料",
                    level_type="零件",
                    lead_time=0,
                    safety_stock=0,
                    lot_size_rule="LFL",
                    is_purchased=True,
                )
                db.add(part)
                db.flush()
                stats["parts"] += 1

            # 创建BOM行：模块→零件（全部放在产品BOM头下，以便递归展开）
            part_qty = float(qty_val) if qty_val and str(qty_val).replace('.', '', 1).isdigit() else 1

            # 检查BOM行是否已存在（用产品BOM头作为统一容器）
            existing_line = db.query(BomLine).filter(
                BomLine.bom_header_id == prod_bom.id,
                BomLine.parent_item_id == mod.id,
                BomLine.item_id == part.id,
            ).first()
            if not existing_line:
                stats["bom_lines"] += 1
                db.add(BomLine(
                    bom_header_id=prod_bom.id,
                    parent_item_id=mod.id,
                    item_id=part.id,
                    quantity=part_qty,
                    level=2,
                    sort_order=stats["bom_lines"],
                ))

    # Step 4: 创建产品→模块的BOM行
    sort_order = 0
    prod_bom = db.query(BomHeader).filter(
        BomHeader.product_id == product.id,
    ).first()
    if not prod_bom:
        prod_bom = BomHeader(
            bom_code=f"BOM-{product_name}",
            product_id=product.id,
            version="A",
            status="生效",
        )
        db.add(prod_bom)
        db.flush()

    for module_name, mod in module_map.items():
        sort_order += 1
        existing = db.query(BomLine).filter(
            BomLine.bom_header_id == prod_bom.id,
            BomLine.parent_item_id == product.id,
            BomLine.item_id == mod.id,
        ).first()
        if not existing:
            db.add(BomLine(
                bom_header_id=prod_bom.id,
                parent_item_id=product.id,
                item_id=mod.id,
                quantity=1,
                level=1,
                sort_order=sort_order,
            ))

    db.commit()
    wb.close()

    return {
        "success": True,
        "message": f"导入完成：{stats['product']}个产品，{stats['modules']}个模块，{stats['parts']}个零件，{stats['bom_lines']}条BOM行",
        "stats": stats,
        "errors": errors[:10],
    }
